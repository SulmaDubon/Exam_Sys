# admin_panel/views.py
import random
from django.forms import ValidationError, formset_factory
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
import openpyxl
import pandas as pd
from users.models import CustomUser
from dashboard_users.models import Examen, Modulo, Pregunta, Respuesta, TipoExamen, UserExam
from dashboard_users.forms import ExamenForm, RespuestaFormSet,  TipoExamenForm, ModuloFormSet, SubirPreguntasForm, PreguntaSimpleForm, PreguntaConEnunciadoForm  # Importar ExamenForm desde admin_panel/forms.py
from users.forms import UserRegistrationForm  # Importar UserRegistrationForm desde users/forms.py
from django.contrib.auth.views import LoginView
from django.contrib import messages 
from datetime import date, datetime, timedelta
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from django.db.models import Max
from django.views.generic.edit import FormView
from django.http import JsonResponse


# from .decorators import es_admin


#-------------------------------------------------------
# Función para verificar si el usuario es administrador
#-------------------------------------------------------
def es_admin(usuario):
    return usuario.is_superuser

#-----------------------------------------------------------
# Login Administrador
#------------------------------------------------------------

#@method_decorator(user_passes_test(es_admin), name='dispatch')
class AdminLoginView(LoginView):
    template_name = 'admin_panel/admin_login.html'

    def get_success_url(self):
        return reverse_lazy('admin_panel:admin_panel')

#-----------------------------------------------------------------
#    ADMIN PANEL
#----------------------------------------------------------------

# Vista principal del panel de administración
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class VistaAdminPanel(View):
    def get(self, request):
        # Renderiza la plantilla principal del panel de administración
        return render(request, 'admin_panel/admin_panel.html')


#-------------------------------------------------------------------
#                 USUARIOS
#--------------------------------------------------------------------

# Vista para listar todos los usuarios
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaUsuarios(ListView):
    model = CustomUser
    template_name = 'admin_panel/lista_usuarios.html'
    context_object_name = 'usuarios'
    # Renderiza la plantilla con la lista de usuarios

# Vista para crear un nuevo usuario
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class CrearUsuario(CreateView):
    model = CustomUser
    form_class = UserRegistrationForm
    template_name = 'admin_panel/formulario_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')
    # Renderiza el formulario para crear un usuario y maneja su creación

# Vista para editar un nuevo usuario
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EditarUsuario(UpdateView):
    model = CustomUser
    form_class = UserRegistrationForm
    template_name = 'admin_panel/formulario_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Usuario actualizado exitosamente.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al actualizar el usuario. Por favor, revisa los datos ingresados.')
        return super().form_invalid(form)

# Vista para eliminar un usuario existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EliminarUsuario(DeleteView):
    model = CustomUser
    template_name = 'admin_panel/confirmar_eliminacion_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')
    # Renderiza una página de confirmación y maneja la eliminación del usuario

#------------------------------------------------------------
#              LISTAR   EXAMEN
#-----------------------------------------------------------

# Vista para listar todos los exámenes
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaExamenes(ListView):
    model = Examen
    template_name = 'admin_panel/lista_examenes.html'
    context_object_name = 'examenes'
    paginate_by = 10  # Número de elementos por página

    def get_queryset(self):
        queryset = Examen.objects.all()
        order = self.request.GET.get('order', 'fecha')
        direction = self.request.GET.get('direction', 'asc')
        year = self.request.GET.get('year')
        month = self.request.GET.get('month')
        
        if year and month:
            queryset = queryset.filter(fecha__year=year, fecha__month=month)
        elif year:
            queryset = queryset.filter(fecha__year=year)
        elif month:
            queryset = queryset.filter(fecha__month=month)
        
        if direction == 'desc':
            order = '-' + order
        
        return queryset.order_by(order)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = datetime.now().year
        current_month = datetime.now().month

        # Obtener valores de los parámetros GET o asignar por defecto el año y mes actuales
        selected_year = self.request.GET.get('year', current_year)
        selected_month = self.request.GET.get('month', current_month)

        # Crear listas de años y meses para el filtrado
        year_list = list(range(current_year - 5, current_year + 5))
        month_list = [
            {'value': i, 'name': datetime(1900, i, 1).strftime('%B')}
            for i in range(1, 13)
        ]

        # Actualizar el contexto con los datos necesarios
        context['year_list'] = year_list
        context['month_list'] = month_list
        context['current_year'] = int(selected_year)  # Asegurarse de que sean enteros
        context['current_month'] = int(selected_month)

        return context

# Vista para crear un nuevo examen
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class CrearExamen(CreateView):
    model = Examen
    form_class = ExamenForm
    template_name = 'admin_panel/formulario_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')

    def form_valid(self, form):
        # Guardar el examen primero
        response = super().form_valid(form)
        examen = self.object

        # Seleccionar preguntas al azar respetando los módulos
        preguntas_seleccionadas = []
        for modulo in examen.tipo_examen.modulos.all():
            preguntas_simples = list(Pregunta.objects.filter(modulo=modulo, enunciado__isnull=True, preguntas_relacionadas__isnull=True))
            enunciados = list(Pregunta.objects.filter(modulo=modulo, enunciado__isnull=True, preguntas_relacionadas__isnull=False).distinct())

            total_preguntas_requeridas = modulo.cantidad_preguntas
            cantidad_anidadas_max = len(enunciados)
            cantidad_simples_max = len(preguntas_simples)

            # Intentar distribuir las preguntas de forma balanceada entre simples y anidadas
            cantidad_anidadas = min(cantidad_anidadas_max, total_preguntas_requeridas // 2)
            cantidad_simples = min(cantidad_simples_max, total_preguntas_requeridas - cantidad_anidadas)

            # Si no hay suficientes preguntas simples, ajustar con preguntas anidadas
            if cantidad_simples < total_preguntas_requeridas // 2:
                cantidad_anidadas = min(cantidad_anidadas_max, total_preguntas_requeridas - cantidad_simples)

            # Si no hay suficientes preguntas anidadas, ajustar con preguntas simples
            if cantidad_anidadas < total_preguntas_requeridas // 2:
                cantidad_simples = min(cantidad_simples_max, total_preguntas_requeridas - cantidad_anidadas)

            # Seleccionar preguntas simples
            if cantidad_simples > 0 and cantidad_simples_max > 0:
                preguntas_simples_seleccionadas = random.sample(preguntas_simples, cantidad_simples)
                preguntas_seleccionadas.extend(preguntas_simples_seleccionadas)

            # Seleccionar preguntas anidadas, asegurando que se incluyan tanto los enunciados como las preguntas relacionadas
            if cantidad_anidadas > 0 and cantidad_anidadas_max > 0:
                enunciados_seleccionados = random.sample(enunciados, cantidad_anidadas)
                for enunciado in enunciados_seleccionados:
                    if enunciado not in preguntas_seleccionadas:
                        preguntas_seleccionadas.append(enunciado)
                        preguntas_relacionadas = list(enunciado.preguntas_relacionadas.all())
                        preguntas_seleccionadas.extend(preguntas_relacionadas[:3])  # Limitar a 3 preguntas relacionadas

        # Asegurar que el número de preguntas seleccionadas no exceda el total requerido por el módulo
        preguntas_seleccionadas_por_modulo = {}
        for pregunta in preguntas_seleccionadas:
            modulo = pregunta.modulo
            if modulo not in preguntas_seleccionadas_por_modulo:
                preguntas_seleccionadas_por_modulo[modulo] = []
            preguntas_seleccionadas_por_modulo[modulo].append(pregunta)

        preguntas_finales = []
        for modulo, preguntas in preguntas_seleccionadas_por_modulo.items():
            preguntas_finales.extend(preguntas[:modulo.cantidad_preguntas])

        # Guardar las preguntas seleccionadas en el examen
        if preguntas_finales:
            examen.preguntas.set(preguntas_finales)
        else:
            messages.error(self.request, 'Insuficientes preguntas para este examen. El examen se ha guardado, pero no tiene suficientes preguntas.')

        messages.success(self.request, f'Se han asignado {len(preguntas_finales)} preguntas al examen.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al crear el examen. Por favor, revisa los datos ingresados.')
        return super().form_invalid(form)


# Vista para editar un examen existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EditarExamen(UpdateView):
    model = Examen
    form_class = ExamenForm
    template_name = 'admin_panel/formulario_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')

# Vista para eliminar un examen existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EliminarExamen(DeleteView):
    model = Examen
    template_name = 'admin_panel/confirmar_eliminacion_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')


#------------------------------------------------------------------
#     USUARIOS POR EXAMEN
#-----------------------------------------------------------------

@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class UsuariosInscritosView(DetailView):
    model = Examen
    template_name = 'admin_panel/usuarios_inscritos.html'
    context_object_name = 'examen'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('search', '')
        
        if search_query:
            usuarios = self.object.usuarios.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(cedula__icontains=search_query)
            )
        else:
            usuarios = self.object.usuarios.all()
        
        paginator = Paginator(usuarios, 20)  # Mostrar 20 usuarios por página
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context['search_query'] = search_query
        context['page_obj'] = page_obj
        return context

#-------------------------------------------------------
#           VISTA PARA MODULOS ASOCIADOS A EXAMEN
#--------------------------------------------------------

def get_modulos(request, tipo_examen_id):
    # Filtra los módulos por el tipo de examen seleccionado
    modulos = Modulo.objects.filter(tipo_examen_id=tipo_examen_id)
    data = {
        'modulos': [{'id': modulo.id, 'nombre': modulo.nombre} for modulo in modulos]
    }
    return JsonResponse(data)


#------------------------------------------------------------------------------
#      PREGUNTAS
#------------------------------------------------------------------------------
# listar pregunta simple
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaPreguntas(ListView):
    model = Pregunta
    template_name = 'admin_panel/lista_preguntas.html'
    context_object_name = 'preguntas'
    paginate_by = 10  # Número de preguntas por página

    def get_queryset(self):
        # Obtener todas las preguntas con sus respuestas relacionadas
        return Pregunta.objects.prefetch_related('respuestas').order_by('id')


@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
# Crear pregunta simple
class PreguntaCreateView(CreateView):
    model = Pregunta
    form_class = PreguntaSimpleForm
    template_name = 'admin_panel/formulario_preguntas.html'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['respuestas'] = RespuestaFormSet(self.request.POST)
        else:
            data['respuestas'] = RespuestaFormSet()
        return data

    def form_valid(self, form):
        # Guardar la pregunta primero para asegurarnos de que tenga un ID
        self.object = form.save()

        # Obtenemos el formset de respuestas
        context = self.get_context_data()
        respuestas = context['respuestas']

        # Asignar la pregunta guardada a las respuestas
        respuestas.instance = self.object

        # Verificar si el formset es válido y guardarlo
        if respuestas.is_valid():
            respuestas.save()
            return redirect(self.success_url)
        else:
            # Si el formset no es válido, renderizar nuevamente el formulario con errores
            return self.render_to_response(self.get_context_data(form=form, respuestas=respuestas))


@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
# Crear pregunta con enunciado
class PreguntaConEnunciadoCreateView(CreateView):
    model = Pregunta
    form_class = PreguntaConEnunciadoForm
    template_name = 'admin_panel/formulario_pregunta_enunciado'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
# Editar pregunta
class PreguntaUpdateView(UpdateView):
    model = Pregunta
    form_class = PreguntaSimpleForm
    template_name = 'admin_panel/formulario_preguntas.html'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['respuestas'] = RespuestaFormSet(self.request.POST, instance=self.object)
        else:
            data['respuestas'] = RespuestaFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        respuestas = context['respuestas']

        # Guardamos primero la pregunta
        self.object = form.save()

        # Luego asignamos la pregunta guardada al formset de respuestas y lo guardamos
        if respuestas.is_valid():
            respuestas.instance = self.object
            respuestas.save()

        return super().form_valid(form)

# Eliminar pregunta
class PreguntaDeleteView(DeleteView):
    model = Pregunta
    template_name = 'admin_panel/confirmar_eliminacion_pregunta.html'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

# subir pregunta
class SubirPreguntasView(FormView):
    template_name = 'admin_panel/subir_preguntas.html'
    form_class = SubirPreguntasForm
    success_url = reverse_lazy('admin_panel:lista_preguntas')

    def form_valid(self, form):
        archivo_excel = form.cleaned_data['archivo']
        tipo_examen = form.cleaned_data['tipo_examen']

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb.active

        modulos_no_existentes = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            _, _, _, _, _, _, modulo_nombre = row
            
            if not Modulo.objects.filter(nombre=modulo_nombre).exists():
                modulos_no_existentes.append(modulo_nombre)
        
        if modulos_no_existentes:
            mensajes_error = f"Los siguientes módulos no existen: {', '.join(modulos_no_existentes)}. No se ha subido el archivo."
            messages.error(self.request, mensajes_error)
            return self.form_invalid(form)

        enunciado_actual = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            texto_enunciado, texto_pregunta, respuesta_correcta, respuesta_1, respuesta_2, respuesta_3, modulo_nombre = row

            modulo = Modulo.objects.get(nombre=modulo_nombre)

            if texto_enunciado:
                enunciado_actual = Pregunta.objects.create(
                    texto=texto_enunciado,
                    modulo=modulo,
                    tipo_examen=tipo_examen,
                    activo=True
                )

            pregunta_actual = Pregunta.objects.create(
                texto=texto_pregunta,
                modulo=modulo,
                tipo_examen=tipo_examen,
                enunciado=enunciado_actual if texto_enunciado else None,
                activo=True
            )

            # Procesar las respuestas y marcar la correcta
            respuestas = [
                {'texto': respuesta_1, 'es_correcta': respuesta_1 == respuesta_correcta},
                {'texto': respuesta_2, 'es_correcta': respuesta_2 == respuesta_correcta},
                {'texto': respuesta_3, 'es_correcta': respuesta_3 == respuesta_correcta},
            ]

            for respuesta in respuestas:
                Respuesta.objects.create(
                    pregunta=pregunta_actual,
                    texto=respuesta['texto'],
                    es_correcta=respuesta['es_correcta']
                )

        return super().form_valid(form)

class ActualizarPreguntasView(View):
    def get(self, request, *args, **kwargs):
        # Ejecutar la función de sincronización manual
        hoy = date.today()
        for examen in Examen.objects.all():
            if examen.fecha < hoy:
                continue

            # Verificar si realmente hay preguntas para sincronizar
            for user_exam in UserExam.objects.filter(examen=examen, finalizado=False):
                if user_exam.preguntas.exists():
                    user_exam.preguntas.set(examen.preguntas.all())
                    user_exam.save()

        # Mensaje de éxito para el usuario solo una vez
        messages.success(request, 'Las preguntas han sido actualizadas correctamente.')

        # Redirigir a la lista de preguntas
        return redirect(reverse('admin_panel:lista_preguntas'))

#------------------------------------------------
#   ACCIONES
#------------------------------------------------
class AccionesExamenesView(View):
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        selected_exams = request.POST.getlist('selected_exams')

        if not selected_exams:
            messages.warning(request, "No se seleccionaron exámenes.")
            return redirect('admin_panel:lista_examenes')

        if action == 'edit':
            if len(selected_exams) > 1:
                messages.warning(request, "Solo puedes editar un examen a la vez.")
                return redirect('admin_panel:lista_examenes')
            return redirect(reverse('admin_panel:editar_examen', args=[selected_exams[0]]))
        
        elif action == 'delete':
            for exam_id in selected_exams:
                exam = get_object_or_404(Examen, id=exam_id)
                exam.delete()
            messages.success(request, "Exámenes eliminados con éxito.")
        
        elif action == 'users':
            if len(selected_exams) > 1:
                messages.warning(request, "Solo puedes ver los usuarios inscritos de un examen a la vez.")
                return redirect('admin_panel:lista_examenes')
            return redirect(reverse('admin_panel:usuarios_inscritos', args=[selected_exams[0]]))

        return redirect('admin_panel:lista_examenes')


#---------------------------------------------
#             CREAR TIPO EXAMEN
#----------------------------------------------

class TipoExamenListView(ListView):
    model = TipoExamen
    template_name = 'admin_panel/lista_tipo_examen.html'
    context_object_name = 'tipos_examenes'

    def get_queryset(self):
        # Utilizamos prefetch_related para traer todos los módulos relacionados en una sola operación
        return TipoExamen.objects.prefetch_related('modulos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Preparar información de módulos para cada tipo de examen
        tipos_examenes_info = []
        for tipo_examen in context['tipos_examenes']:
            # `tipo_examen.modulos.all()` usa los módulos prefetched gracias a `prefetch_related`
            modulos = tipo_examen.modulos.all()
            total_preguntas = sum(modulo.cantidad_preguntas for modulo in modulos)
            
            tipos_examenes_info.append({
                'tipo_examen': tipo_examen,
                'modulos': modulos,
                'total_preguntas': total_preguntas
            })

        # Añadir la información enriquecida al contexto
        context['tipos_examenes_info'] = tipos_examenes_info
        return context



class CrearTipoExamenView(View):
    def get(self, request):
        tipo_examen_form = TipoExamenForm()
        modulo_formset = ModuloFormSet()
        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'admin_panel/crear_tipo_examen.html', context)

    def post(self, request):
        tipo_examen_form = TipoExamenForm(request.POST)
        modulo_formset = ModuloFormSet(request.POST)

        if tipo_examen_form.is_valid() and modulo_formset.is_valid():
            tipo_examen = tipo_examen_form.save()
            
             # Vincular los módulos al tipo de examen
            modulo_formset.instance = tipo_examen
            modulo_formset.save()
            return redirect('admin_panel:lista_tipo_examen')  # Cambia 'ruta_donde_redirigir' a la URL de listar

        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'admin_panel/crear_tipo_examen.html', context)




class EditarTipoExamenView(View):
    def get(self, request, pk):
        tipo_examen = get_object_or_404(TipoExamen, pk=pk)
        tipo_examen_form = TipoExamenForm(instance=tipo_examen)
        modulo_formset = ModuloFormSet(instance=tipo_examen)
        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'editar_tipo_examen.html', context)

    def post(self, request, pk):
        tipo_examen = get_object_or_404(TipoExamen, pk=pk)
        tipo_examen_form = TipoExamenForm(request.POST, instance=tipo_examen)
        modulo_formset = ModuloFormSet(request.POST, instance=tipo_examen)

        if tipo_examen_form.is_valid() and modulo_formset.is_valid():
            tipo_examen = tipo_examen_form.save()
            modulo_formset.save()
            return redirect('ruta_donde_redirigir')

        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'editar_tipo_examen.html', context)
    

class TipoExamenDeleteView(DeleteView):
    model = TipoExamen
    template_name = 'admin_panel/confirmar_eliminar_tipo_examen.html'
    success_url = reverse_lazy('admin_panel:listar_tipo_examen')  # Redirige a la lista de exámenes después de eliminar


